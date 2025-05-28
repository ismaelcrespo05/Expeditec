from django.shortcuts import render
from django.views import View
from Login import views as Login_views
from Administrador import models as Admin_models,utils as Admin_utils
from django.http import HttpRequest

from Chat import models as Chat_models
from django.http import JsonResponse
from django.db.models import Q
from datetime import datetime
from Aspirante import models as Aspirante_models
from RRHH import models as RRHH_models


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.db.models import Prefetch

# Create your views here.
from django.db.models.functions import Concat
from django.db.models import Value, CharField
from django.db.models import F
from Notificacion import views as Notificacion_views


class RRHH_Dashboard(View):
    def get(self, request):
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
            cant_profesores = Admin_models.Aspirante.objects.filter(tipo='Profesor',userid__is_active=True).count()
            cant_estudiantes = Admin_models.Aspirante.objects.filter(tipo='Estudiante',userid__is_active=True).count()
            cantidad_x_categoria = Admin_utils.contar_por_categoria_docente()
            total_usuarios = cant_estudiantes + cant_profesores

            distribucion_x_categoria = {}
            for categoria, cantidad in cantidad_x_categoria.items():
                distribucion_x_categoria[categoria] = round((cantidad / total_usuarios) * 100, 2) if total_usuarios != 0 else 0

            # Contar solicitudes por estado
            from Aspirante import models as Aspirante_models
            solicitudes_pendientes = Aspirante_models.SolicitudCambioCategoria.objects.filter(estado='Pendiente',aspirante__userid__is_active=True).count()
            solicitudes_en_revision = Aspirante_models.SolicitudCambioCategoria.objects.filter(estado='En revisión',aspirante__userid__is_active=True).count()

            return render(request, 'RRHH/dashboard.html', {
                'rrhh': rrhh,
                'Dashboard': True,
                'cant_profesores': cant_profesores,
                'cant_estudiantes': cant_estudiantes,
                'total_personal': total_usuarios,
                'cantidad_x_categoria': cantidad_x_categoria,
                'distribucion_x_categoria': distribucion_x_categoria,
                'solicitudes_pendientes': solicitudes_pendientes,
                'solicitudes_en_revision': solicitudes_en_revision,
            })

        except Exception as e:
            return Login_views.redirigir_usuario(request=request)

    def post(self,request):
        pass






class Solicitudes(View):
    @staticmethod
    def Notificacion(request, Error=None, Success=None):
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
            
            solicitudes = Aspirante_models.SolicitudCambioCategoria.objects.filter(aspirante__userid__is_active=True).order_by('-fecha_solicitud')
            
            ESTADOS = Aspirante_models.ESTADOS
            CATEGORIAS = Admin_models.CATEGORIA_DOCENTE_CHOICES
            
            solicitudes_por_estado = {
                estado: solicitudes.filter(estado=estado) 
                for estado in ESTADOS
            }
            
            total_solicitudes = solicitudes.count()
            tiene_solicitudes_activas = solicitudes.filter(estado__in=['Pendiente', 'En revisión']).exists()

            return render(request, 'RRHH/solicitudes.html', {
                "Cambio_categoria": True,
                "solicitudes_por_estado": solicitudes_por_estado,
                "estados_filtro": ESTADOS,
                "categorias_filtro": CATEGORIAS,
                "total_solicitudes": total_solicitudes,
                "puede_solicitar": not tiene_solicitudes_activas,
                'rrhh': rrhh,
                'Solicitudes': True,
                'CARGOS_CHOICES': ['Presidente','Secretario'],
                'Error':Error,'Success':Success
            })
        except Exception as e:
            print(e)
            return Login_views.redirigir_usuario(request=request)

    def get(self, request: HttpRequest):
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
            
            solicitudes = Aspirante_models.SolicitudCambioCategoria.objects.filter(aspirante__userid__is_active=True).order_by('-fecha_solicitud')
            
            ESTADOS = Aspirante_models.ESTADOS
            CATEGORIAS = Admin_models.CATEGORIA_DOCENTE_CHOICES

            solicitudes_por_estado = {
                estado: solicitudes.filter(estado=estado) 
                for estado in ESTADOS
            }
            
            total_solicitudes = solicitudes.count()
            tiene_solicitudes_activas = solicitudes.filter(estado__in=['Pendiente', 'En revisión']).exists()

            return render(request, 'RRHH/solicitudes.html', {
                "Cambio_categoria": True,
                "solicitudes_por_estado": solicitudes_por_estado,
                "estados_filtro": ESTADOS,
                "categorias_filtro": CATEGORIAS,
                "total_solicitudes": total_solicitudes,
                "puede_solicitar": not tiene_solicitudes_activas,
                'rrhh': rrhh,
                'Solicitudes': True,
                'CARGOS_CHOICES': ['Presidente','Secretario'],
            })
        
        except Exception as e:
            print(e)
            return Login_views.redirigir_usuario(request=request)




class Rechazar_Solicitud(View):
    def get(self, request):
        return Solicitudes.Notificacion(request=request)
    
    def post(self, request:HttpRequest):
        rrhh = None
        solicitud_id = None
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
        except Exception as e:
            return Login_views.redirigir_usuario(request=request)
        
        try:
            solicitud_id = request.POST.get('solicitud_id')
            solicitud = Aspirante_models.SolicitudCambioCategoria.objects.get(id=solicitud_id,estado="Pendiente",aspirante__userid__is_active=True)
            solicitud.estado = 'Rechazada'
            if not request.POST.get('observaciones') in [None, '']:
                solicitud.observaciones = request.POST.get('observaciones')
            solicitud.save()
            return Solicitudes.Notificacion(request=request, Success="Solicitud rechazada con éxito")
        except Exception as e:
            return Solicitudes.Notificacion(request=request, Error="Solicitud no encontrada")




from datetime import datetime
from django.http import HttpRequest
from django.shortcuts import render
from django.db.models import Q, F, Value, CharField
from django.db.models.functions import Concat

def filtrar_solicitudes(request: HttpRequest):
    if not request.method == 'POST':
        return Solicitudes.Notificacion(request=request)

    # Query inicial
    solicitudes = Aspirante_models.SolicitudCambioCategoria.objects.filter(aspirante__userid__is_active=True).order_by('-fecha_solicitud')

    # Opciones de filtro disponibles
    ESTADOS = Aspirante_models.ESTADOS
    CATEGORIAS = Admin_models.CATEGORIA_DOCENTE_CHOICES

    # Recogida de filtros
    filtros = {
        'estado': request.POST.get('estado'),
        'categoria': request.POST.get('categoria'),
        'fecha_inicio': request.POST.get('fecha_inicio'),
        'fecha_fin': request.POST.get('fecha_fin'),
        'area': request.POST.get('area'),
        'especialidad_solicitada': request.POST.get('especialidad_solicitada'),
        'busqueda': request.POST.get('busqueda'),
    }

    # Aplicación de filtros
    if filtros['estado'] and filtros['estado'] != 'Todos':
        solicitudes = solicitudes.filter(estado=filtros['estado'])

    if filtros['categoria'] and filtros['categoria'] != 'Todas':
        solicitudes = solicitudes.filter(categoria_solicitada=filtros['categoria'])

    if filtros['fecha_inicio']:
        try:
            fecha_inicio = datetime.strptime(filtros['fecha_inicio'], '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_inicio)
        except (ValueError, TypeError):
            pass

    if filtros['fecha_fin']:
        try:
            fecha_fin = datetime.strptime(filtros['fecha_fin'], '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_fin)
        except (ValueError, TypeError):
            pass

    if filtros['area']:
        solicitudes = solicitudes.filter(area__icontains=filtros['area'])

    if filtros['especialidad_solicitada']:
        solicitudes = solicitudes.filter(
            especialidad_solicitada__icontains=filtros['especialidad_solicitada']
        )

    if filtros['busqueda']:
        # Creamos la expresión para nombre completo
        nombre_completo = Concat(
            F('aspirante__nombres'), Value(' '),
            F('aspirante__primer_apellido'), Value(' '),
            F('aspirante__segundo_apellido'),
            output_field=CharField()
        )
        solicitudes = solicitudes.annotate(nombre_completo=nombre_completo).filter(
            Q(aspirante__nombres__icontains=filtros['busqueda']) |
            Q(aspirante__primer_apellido__icontains=filtros['busqueda']) |
            Q(aspirante__segundo_apellido__icontains=filtros['busqueda']) |
            Q(aspirante__ci__icontains=filtros['busqueda']) |
            Q(aspirante__solapin__icontains=filtros['busqueda']) |
            Q(nombre_completo__icontains=filtros['busqueda'])
        )

    # Agrupar por estado para la vista
    solicitudes_por_estado = {
        estado: solicitudes.filter(estado=estado)
        for estado in ESTADOS
    }

    # Datos extra para el contexto
    try:
        rrhh = Admin_models.RRHH.objects.get(userid=request.user)
    except Admin_models.RRHH.DoesNotExist:
        rrhh = None

    tiene_solicitudes_activas = solicitudes.filter(
        estado__in=['Pendiente', 'En revisión']
    ).exists()

    context = {
        "Cambio_categoria": True,
        "solicitudes_por_estado": solicitudes_por_estado,
        "estados_filtro": ESTADOS,
        "categorias_filtro": CATEGORIAS,
        "filtros_aplicados": filtros,
        "total_solicitudes": solicitudes.count(),
        "puede_solicitar": not tiene_solicitudes_activas,
        "rrhh": rrhh,
        "Solicitudes": True,
        "CARGOS_CHOICES": ['Presidente', 'Secretario'],
    }

    return render(request, 'RRHH/solicitudes.html', context)

#################################################################################################################
####################        TRIBUNAL       ######################################################################
#################################################################################################################


class Asignar_tribunal(View):
    def get(self, request):
        return Solicitudes.Notificacion(request=request)

    def post(self, request: HttpRequest):
        if request.method == "POST":
            try:
                rrhh = Admin_models.RRHH.objects.get(userid=request.user)
            except Exception as e:
                print(e)
                return Login_views.redirigir_usuario(request=request)
            
            solicitud_id = request.POST.get('solicitud_id')
            solicitud = None
            try:
                solicitud = Aspirante_models.SolicitudCambioCategoria.objects.get(id=solicitud_id)
            except Exception as e:
                print(e)
                return Solicitudes.Notificacion(request=request, Error="La solicitud no está registrada.")
            
            if solicitud.estado != 'Pendiente':
                return Solicitudes.Notificacion(request=request, Error="La solicitud debe estar en estado Pendiente.")
            
            profesor_ci = str(request.POST.get('profesor_ci')).strip()
            profesor = None
            try:
                profesor = Admin_models.Aspirante.objects.get(ci=profesor_ci, tipo="Profesor",userid__is_active=True)
            except Exception as e:
                print(e)
                return Solicitudes.Notificacion(request=request, Error="El profesor no está registrado.")
            
            # Verificar que el profesor no sea el mismo que hizo la solicitud
            if profesor.userid == solicitud.aspirante.userid:
                return Solicitudes.Notificacion(request=request, Error="No puede asignar al propio solicitante como tribunal.")
            
            cargo = request.POST.get('cargo')
            if cargo not in ['Presidente', 'Secretario']:
                return Solicitudes.Notificacion(request=request, Error="Seleccione un cargo válido.")
            
            # Diccionario de validación de categorías según Resolución 145/2023
            # Clave: categoría solicitada
            # Valor: lista de categorías permitidas para los miembros del tribunal
            CATEGORIAS_TRIBUNAL = {
                'Profesor Titular': ['Profesor Titular'],
                'Profesor Auxiliar': ['Profesor Titular'],
                'Profesor Asistente': ['Profesor Titular', 'Profesor Auxiliar'],
                'Instructor': ['Profesor Titular', 'Profesor Auxiliar'],
                'ATD Superior': ['Profesor Titular', 'Profesor Auxiliar'],
                'ATD Medio Superior': ['Profesor Auxiliar', 'Profesor Asistente']
            }
            
            # Validación de categorías
            categoria_solicitada = solicitud.categoria_solicitada
            categoria_profesor = profesor.categoria_docente
            
            if categoria_solicitada not in CATEGORIAS_TRIBUNAL:
                return Solicitudes.Notificacion(
                    request=request,
                    Error=f"Categoría solicitada {categoria_solicitada} no tiene una configuración de tribunal definida."
                )
            
            if categoria_profesor not in CATEGORIAS_TRIBUNAL[categoria_solicitada]:
                categorias_permitidas = ", ".join(CATEGORIAS_TRIBUNAL[categoria_solicitada])
                return Solicitudes.Notificacion(
                    request=request,
                    Error=f"Para evaluar {categoria_solicitada}, el tribunal debe estar compuesto por Profesores con categorías: {categorias_permitidas}."
                )
            
            tribunal, created = RRHH_models.Tribunal.objects.get_or_create(solicitud_id=solicitud)
            
            # Validar que no exista ya un miembro con el mismo cargo en el tribunal
            if RRHH_models.Miembro_tribunal.objects.filter(tribunal_id=tribunal, cargo=cargo).exists():
                return Solicitudes.Notificacion(
                    request=request, 
                    Error=f"Ya existe un {cargo} asignado a este tribunal. Solo puede haber un {cargo} por tribunal."
                )
            
            # Validar que el profesor no esté ya asignado al tribunal (en cualquier cargo)
            if RRHH_models.Miembro_tribunal.objects.filter(tribunal_id=tribunal, miembro=profesor).exists():
                return Solicitudes.Notificacion(request=request, Error="El profesor ya fue asignado a este tribunal.")
            
            chat = Chat_models.Chat.objects.get(solicitud_id=solicitud)
            Chat_models.Miembro_Chat.objects.create(
                chat_id=chat, userid=profesor.userid,
                tipo="Tribunal"
            )

            miembro = RRHH_models.Miembro_tribunal(tribunal_id=tribunal, miembro=profesor, cargo=cargo)
            miembro.save()
            
            return Solicitudes.Notificacion(request=request, Success=f"El profesor fue asignado como {cargo} con éxito.")
        else:
            return Solicitudes.Notificacion(request=request)


class EliminarMiembroTribunal(View):
    def get(self,request):
        return Solicitudes.Notificacion(request=request)
    
    def post(self, request):
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
        except Exception as e:
            return Login_views.redirigir_usuario(request=request)
        
        try:
            solicitud_id = request.POST.get('solicitud_id')
            member_id = request.POST.get('member_id')
            
            solicitud = Aspirante_models.SolicitudCambioCategoria.objects.get(id=solicitud_id)
            if solicitud.estado != 'Pendiente':
                return Solicitudes.Notificacion(request=request, Error="Solo se pueden modificar tribunales en solicitudes pendientes")
            
            tribunal = RRHH_models.Tribunal.objects.get(solicitud_id=solicitud)
            miembro = RRHH_models.Miembro_tribunal.objects.get(id=member_id, tribunal_id=tribunal)
            chat = Chat_models.Chat.objects.get(solicitud_id=solicitud)
            miembro_chat = Chat_models.Miembro_Chat.objects.get(chat_id=chat,userid=miembro.miembro.userid)
            miembro_chat.delete()
            miembro.delete()
            
            # Check if there are any members left in the tribunal using the related_name
            if not tribunal.miembros.exists():  # Using the related_name 'miembros'
                tribunal.delete()
            
            return Solicitudes.Notificacion(request=request, Success="El profesor ha sido eliminado del tribunal correctamente.")
            
        except Aspirante_models.SolicitudCambioCategoria.DoesNotExist:
            return Solicitudes.Notificacion(request=request, Error="Solicitud no encontrada")
        except RRHH_models.Tribunal.DoesNotExist:
            return Solicitudes.Notificacion(request=request, Error="Tribunal no encontrado")
        except RRHH_models.Miembro_tribunal.DoesNotExist:
            return Solicitudes.Notificacion(request=request, Error="Miembro no encontrado")
        except Exception as e:
            return Solicitudes.Notificacion(request=request, Error=str(e))    






class Pasar_a_revision(View):
    def get(self, request):
        return Solicitudes.Notificacion(request=request)
    
    def post(self,request):
        try:
            rrhh = Admin_models.RRHH.objects.get(userid=request.user)
        except Exception as e:
            return Login_views.redirigir_usuario(request=request)
        
        try:
            solicitud_id = request.POST.get('solicitud_id')
            solicitud = Aspirante_models.SolicitudCambioCategoria.objects.get(id=solicitud_id)
            if solicitud.estado != 'Pendiente':
                return Solicitudes.Notificacion(request=request,Error="Solo se pueden pasar a revision solicitudes pendientes")
            
            solicitud.estado = 'En revisión'
            solicitud.save()
            Notificacion_views.get_mail(solicitud=solicitud,tipo="EN_REVISION")
            return Solicitudes.Notificacion(request=request,Success="La solicitud ha sido pasada a revision correctamente.")
        except Exception as e:
            return Solicitudes.Notificacion(request=request,Error=str(e))




def exportar_solicitudes(request: HttpRequest):
    if not request.method == 'POST':
        return Solicitudes.Notificacion(request=request)

    # Prefetch para optimizar las consultas
    tribunal_prefetch = Prefetch(
        'tribunal',
        queryset=RRHH_models.Tribunal.objects.select_related('solicitud_id').prefetch_related(
            Prefetch(
                'miembros',
                queryset=RRHH_models.Miembro_tribunal.objects.select_related('miembro', 'miembro__userid'),
                to_attr='miembros_tribunal'
            )
        ),
        to_attr='tribunal_data'
    )

    solicitudes = Aspirante_models.SolicitudCambioCategoria.objects.all()\
        .order_by('-fecha_solicitud')\
        .select_related('aspirante', 'aspirante__userid')\
        .prefetch_related(tribunal_prefetch)
    
    # Aplicar filtros
    ESTADOS = Aspirante_models.ESTADOS
    CATEGORIAS = Admin_models.CATEGORIA_DOCENTE_CHOICES
    
    filtros = {
        'estado': request.POST.get('estado'),
        'categoria': request.POST.get('categoria'),
        'fecha_inicio': request.POST.get('fecha_inicio'),
        'fecha_fin': request.POST.get('fecha_fin'),
        'busqueda': request.POST.get('busqueda'),
        'area': request.POST.get('area'),
        'especialidad_solicitada': request.POST.get('especialidad_solicitada')
    }

    if filtros['estado'] and filtros['estado'] != 'Todos':
        solicitudes = solicitudes.filter(estado=filtros['estado'])
    
    if filtros['categoria'] and filtros['categoria'] != 'Todas':
        solicitudes = solicitudes.filter(categoria_solicitada=filtros['categoria'])
    
    if filtros['fecha_inicio']:
        try:
            fecha_inicio = datetime.strptime(filtros['fecha_inicio'], '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__gte=fecha_inicio)
        except (ValueError, TypeError):
            pass
    
    if filtros['fecha_fin']:
        try:
            fecha_fin = datetime.strptime(filtros['fecha_fin'], '%Y-%m-%d').date()
            solicitudes = solicitudes.filter(fecha_solicitud__date__lte=fecha_fin)
        except (ValueError, TypeError):
            pass
    
    if filtros['area']:
        solicitudes = solicitudes.filter(area__icontains=filtros['area'])
    
    if filtros['especialidad_solicitada']:
        solicitudes = solicitudes.filter(especialidad_solicitada__icontains=filtros['especialidad_solicitada'])
    
    if filtros['busqueda']:
        # Creamos una expresión para el nombre completo
        nombre_completo_expr = Concat(
            F('aspirante__nombres'), Value(' '),
            F('aspirante__primer_apellido'), Value(' '),
            F('aspirante__segundo_apellido'),
            output_field=CharField()
        )
        
        solicitudes = solicitudes.annotate(
            nombre_completo=nombre_completo_expr
        ).filter(
            Q(aspirante__nombres__icontains=filtros['busqueda']) |
            Q(aspirante__primer_apellido__icontains=filtros['busqueda']) |
            Q(aspirante__segundo_apellido__icontains=filtros['busqueda']) |
            Q(aspirante__ci__icontains=filtros['busqueda']) |
            Q(aspirante__solapin__icontains=filtros['busqueda']) |
            Q(nombre_completo__icontains=filtros['busqueda'])
        )

    # Resto del código permanece igual...
    # Crear el libro de Excel
    wb = Workbook()
    ws_solicitudes = wb.active
    ws_solicitudes.title = "Solicitudes"
    
    # Hoja para miembros de tribunal
    ws_tribunales = wb.create_sheet(title="Tribunales")

    # Encabezados para solicitudes
    headers_solicitudes = [
        'ID Solicitud', 'CI Aspirante', 'Nombres', 'Apellidos', 'Email',
        'Categoría Actual', 'Grado Científico', 'Especialidad',
        'Categoría Solicitada', 'Área', 'Cargo Actual', 
        'Especialidad Solicitada', 'Estado', 'Fecha Solicitud', 
        'Fecha Resolución', 'Observaciones',
        'Presidente CI', 'Presidente Nombre', 'Presidente Categoría',
        'Secretario CI', 'Secretario Nombre', 'Secretario Categoría',
        'Suplente 1 CI', 'Suplente 1 Nombre', 'Suplente 1 Categoría',
        'Suplente 2 CI', 'Suplente 2 Nombre', 'Suplente 2 Categoría',
        'Miembro 1 CI', 'Miembro 1 Nombre', 'Miembro 1 Categoría',
        'Miembro 2 CI', 'Miembro 2 Nombre', 'Miembro 2 Categoría',
        'Miembro 3 CI', 'Miembro 3 Nombre', 'Miembro 3 Categoría'
    ]
    
    ws_solicitudes.append(headers_solicitudes)
    
    # Encabezados para tribunales
    headers_tribunales = [
        'ID Solicitud', 'CI Aspirante', 'Nombres Aspirante', 
        'Cargo en Tribunal', 'CI Miembro', 'Nombres Miembro',
        'Apellidos Miembro', 'Email Miembro', 'Categoría Miembro',
        'Especialidad Miembro', 'Grado Científico Miembro'
    ]
    ws_tribunales.append(headers_tribunales)
    
    # Estilo para los encabezados
    for sheet in [ws_solicitudes, ws_tribunales]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Función para obtener datos formateados de un miembro
    def datos_miembro(miembro):
        if miembro and hasattr(miembro, 'miembro'):
            return [
                miembro.miembro.ci,
                miembro.miembro.nombres,
                f"{miembro.miembro.primer_apellido} {miembro.miembro.segundo_apellido or ''}".strip(),
                miembro.miembro.userid.email if miembro.miembro.userid else '',
                miembro.miembro.categoria_docente or "",
                miembro.miembro.especialidad or "",
                miembro.miembro.grado_cientifico or ""
            ]
        return ["", "", "", "", "", "", ""]

    # Llenar datos de solicitudes y tribunales
    for solicitud in solicitudes:
        aspirante = solicitud.aspirante
        tribunal = getattr(solicitud, 'tribunal_data', None)
        miembros_tribunal = getattr(tribunal, 'miembros_tribunal', []) if tribunal else []
        
        # Obtener miembros específicos
        presidente = next((m for m in miembros_tribunal if m.cargo == "Presidente"), None)
        secretario = next((m for m in miembros_tribunal if m.cargo == "Secretario"), None)
        suplentes = [m for m in miembros_tribunal if m.cargo == "Suplente"]
        miembros = [m for m in miembros_tribunal if m.cargo == "Miembro"]
        
        # Asegurar que hay al menos 2 suplentes y 3 miembros (rellenar con vacíos si no existen)
        suplente1 = suplentes[0] if len(suplentes) > 0 else None
        suplente2 = suplentes[1] if len(suplentes) > 1 else None
        miembro1 = miembros[0] if len(miembros) > 0 else None
        miembro2 = miembros[1] if len(miembros) > 1 else None
        miembro3 = miembros[2] if len(miembros) > 2 else None

        # Datos básicos de la solicitud
        row_data = [
            solicitud.id,
            aspirante.ci,
            aspirante.nombres,
            f"{aspirante.primer_apellido} {aspirante.segundo_apellido or ''}".strip(),
            aspirante.userid.email if aspirante.userid else '',
            aspirante.categoria_docente,
            aspirante.grado_cientifico,
            aspirante.especialidad,
            solicitud.categoria_solicitada,
            solicitud.area,
            solicitud.cargo_actual,
            solicitud.especialidad_solicitada,
            solicitud.estado,
            solicitud.fecha_solicitud.strftime('%Y-%m-%d %H:%M:%S'),
            solicitud.fecha_resolucion.strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha_resolucion else '',
            solicitud.observaciones or ''
        ]
        
        # Agregar datos de cada miembro del tribunal
        for miembro in [presidente, secretario, suplente1, suplente2, miembro1, miembro2, miembro3]:
            datos = datos_miembro(miembro)
            row_data.extend([datos[0], f"{datos[1]} {datos[2]}", datos[4]])
        
        ws_solicitudes.append(row_data)
        
        # Datos detallados para la hoja de tribunales
        for miembro in miembros_tribunal:
            miembro_data = datos_miembro(miembro)
            ws_tribunales.append([
                solicitud.id,
                aspirante.ci,
                f"{aspirante.nombres} {aspirante.primer_apellido}",
                miembro.cargo,
                miembro_data[0],  # CI
                miembro_data[1],  # Nombres
                miembro_data[2],  # Apellidos
                miembro_data[3],  # Email
                miembro_data[4],  # Categoría
                miembro_data[5],  # Especialidad
                miembro_data[6]   # Grado científico
            ])

    # Ajustar el ancho de las columnas
    for sheet in [ws_solicitudes, ws_tribunales]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Preparar la respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=solicitudes_y_tribunales.xlsx'
    wb.save(response)

    return response



